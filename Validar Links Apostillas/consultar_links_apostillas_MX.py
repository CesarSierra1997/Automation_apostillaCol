#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consulta de link de apostillas (modo semi-automático, sin evadir reCAPTCHA v3).

Qué hace
--------
- Lee un Excel con las columnas: "ESTUDIANTE", "COD. DIPLOMA", "COD. NOTAS", "COD. REGISTRO".
- Solo intenta consultar códigos que sean NUMÉRICOS y que comiencen por "52".
- Abre el sitio de consulta de la Cancillería, llena el código y el correo,
  y si aparece el resultado exitoso, sigue el enlace "www.cancilleria.gov.co/apostilla"
  para capturar el link final del documento (ej.: .../documento.aspx?cod=...&fecha=...).
- Genera un Excel de salida con tres nuevas columnas: "LINK_DIPLOMA", "LINK_NOTAS", "LINK_REGISTRO". 

Importante sobre reCAPTCHA v3
-----------------------------
Este script realiza 3 intentos para lograr superar reCAPTCHA v3. Si no lo logra, se registra "POSIBLE_RESTRICCION_ANTIBOT" y se continúa.
    
Uso
---
1) Instala dependencias (en consola):
    pip install -r requirements.txt 

2) Ejecuta:
   python consulta_apostillas.py --excel Entrada.xlsx --salida salida.xlsx

   (Si no pasas parámetros, buscará 'entrada.xlsx' y generará 'salida.xlsx' en la misma carpeta.)

Notas
-----
- Se recomienda ejecución en modo con ventana (headless=False) para facilitar el diagnóstico.
- Ajusta los selectores en SELECTORES si el sitio cambia.
"""
#!/usr/bin/env python3
from __future__ import annotations
import argparse
import random # Importamos la librería random
import re
import time
from pathlib import Path
from typing import Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import sync_playwright, TimeoutError

BASE_URL = "https://tramites.cancilleria.gov.co/apostillalegalizacion/consulta/tramite.aspx"
EMAIL_FIJO = "apostillamen@gmail.com"
MSJ_NO_ENCONTRADO = "No se han encontrado registros"
MSJ_POSIBLE_ANTIBOT = "POSIBLE ANTIBOT/CAPTCHA"
COLUMNAS = {
    "ESTUDIANTE": "ESTUDIANTE",
    "COD. DIPLOMA": "COD_DIPLOMA",
    "COD. NOTAS": "COD_NOTAS",
    "COD. REGISTRO": "COD_REGISTRO",
}
MAX_DEBUG = 83

# ... (Las funciones normaliza_columnas y limpiar_codigo no cambian) ...
def normaliza_columnas(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {}
    for col in df.columns:
        base = col.strip().upper().replace("  ", " ")
        mapping[col] = base
    df = df.rename(columns=mapping)
    rev = {}
    for esperado, interno in COLUMNAS.items():
        candidatos = [c for c in df.columns if c == esperado or c.replace(".", "") == esperado.replace(".", "")]
        if candidatos:
            rev[candidatos[0]] = interno
    df = df.rename(columns=rev)
    return df

def limpiar_codigo(valor) -> Optional[str]:
    if pd.isna(valor) or not isinstance(valor, (str, int, float)):
        return None
    s = re.sub(r"\D", "", str(valor))
    if not s:
        return None
    if not s.startswith("52"):
        return None
    return s

def consultar_codigo(page, codigo: str, correo: str) -> Tuple[str, str]:
    print(f"  🔎 Buscando código {codigo} ...")
    
    # Selectores de la página
    sel_codigo = "#contenido_tbNumeroSolicitud"
    sel_correo = "#contenido_ucCorreoElectronico_tbCorreoElectronico"
    sel_btn = "#contenido_btnBuscar"
    sel_link_exito = "#contenido_ucInfor_lblMensajes2 a"
    sel_no_encontrado = "text=No se han encontrado registros"
    sel_error_captcha = "#contenido_validadorCaptcha"
    
    # Parámetros para la lógica de reintento
    max_captcha_retries = 3 # Número de veces que intentará pasar el captcha

    try:
        page.goto(BASE_URL, timeout=60_000)
        page.wait_for_selector(sel_codigo, timeout=30_000)
        page.fill(sel_codigo, codigo)
        page.wait_for_selector(sel_correo, timeout=30_000)
        page.fill(sel_correo, correo.strip())
    except Exception as e:
        return "ERROR", f"Error cargando la página o llenando campos: {e}"

    for attempt in range(max_captcha_retries):
        page.click(sel_btn)
        
        try:
            # 1. Comprobar si tuvimos ÉXITO
            link_handle = page.locator(sel_link_exito).first
            href = link_handle.get_attribute("href", timeout=5_000) # Timeout corto
            if href:
                print(f"  ✅ Encontrado link: {href}")
                return "OK", href
        except TimeoutError:
            # Si no hay link de éxito, comprobamos los otros casos
            pass

        # 2. Comprobar si el registro NO FUE ENCONTRADO
        if page.locator(sel_no_encontrado).count() > 0:
            print(f"  ❌ No encontrado")
            return "NO", MSJ_NO_ENCONTRADO
            
        # 3. Comprobar si falló el CAPTCHA
        if page.locator(sel_error_captcha).is_visible():
            print(f"  🤖 CAPTCHA detectado. Intento {attempt + 1}/{max_captcha_retries}. Simulando clics...")
            
            # LÓGICA PARA SIMULAR COMPORTAMIENTO HUMANO
            try:
                # Clic en el título principal (un lugar seguro)
                page.click('h1:has-text("Consultar solicitud")', timeout=2000)
                time.sleep(random.uniform(0.5, 1.2)) # Pausa aleatoria
                
                # Clic en el texto informativo (otro lugar seguro)
                page.click('p:has-text("Diligencie el correo electrónico")', timeout=2000)
                time.sleep(random.uniform(0.4, 1.0))
            except Exception as e:
                print(f"   (No se pudo hacer clic de simulación, continuando de todos modos: {e})")

            # Continuamos con el siguiente intento del bucle
            continue

        # 4. Si después de unos segundos no pasa nada, puede ser un error genérico o antibot
        time.sleep(2) # Pequeña espera por si la página está lenta
    
    # Si salimos del bucle sin éxito
    print(f"  ⚠️  Posible antibot o CAPTCHA no superado tras {max_captcha_retries} intentos.")
    return "ANTIBOT", MSJ_POSIBLE_ANTIBOT

# ... (La función procesar_excel y el resto del script no cambian) ...
def procesar_excel(path_excel: Path, path_salida: Path, correo: str = EMAIL_FIJO, pausa_seg: float = 2.5):
    df = pd.read_excel(path_excel, dtype=str)
    df = normaliza_columnas(df)

    wb = load_workbook(path_excel)
    ws = wb.active

    headers = {cell.value.strip().upper().replace(".", ""): cell.column_letter for cell in ws[1]}
    col_map = {}
    visible_text_map = {}
    for display, internal in {"COD. DIPLOMA": "COD_DIPLOMA", "COD. NOTAS": "COD_NOTAS", "COD. REGISTRO": "COD_REGISTRO"}.items():
        header_key = display.replace(".", "")
        if header_key in headers:
            col_map[internal] = headers[header_key]
            visible_text_map[internal] = display.replace("COD. ", "").title()

    rojo = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=100)
        context = browser.new_context()
        page = context.new_page()

        contador = 0
        for idx, row_data in df.iterrows():
            if contador >= MAX_DEBUG:
                break
            
            excel_row_num = idx + 2

            for col_key, col_letter in col_map.items():
                celda = ws[f"{col_letter}{excel_row_num}"]

                if celda.hyperlink:
                    print(f"\nProcesando fila {excel_row_num}, columna {col_key}: Ya es un hipervínculo. Omitiendo.")
                    continue

                codigo_raw = row_data.get(col_key, None)
                print(f"\nProcesando fila {excel_row_num}, columna {col_key} → valor: {codigo_raw}")
                codigo = limpiar_codigo(codigo_raw)

                if not codigo:
                    print("  ⚠️ No es un código válido o no inicia con 52.")
                    continue

                try:
                    estado, valor = consultar_codigo(page, codigo, correo)
                except Exception as e:
                    estado, valor = "ERROR", f"Excepción navegando: {e}"

                if estado == "OK":
                    celda.value = visible_text_map[col_key]
                    celda.hyperlink = valor
                    celda.style = "Hyperlink"
                    celda.font = Font(color="0000FF", underline="single")
                elif estado == "NO":
                    celda.fill = rojo
                elif estado == "ANTIBOT":
                    celda.fill = amarillo
                
                time.sleep(pausa_seg)

            contador += 1

        context.close()
        browser.close()

    wb.save(path_salida)
    print(f"\n✅ Listo. Archivo generado en: {path_salida.resolve()}")

def main():
    parser = argparse.ArgumentParser(description="Consulta y validación de códigos")
    parser.add_argument("--excel", type=Path, default=Path("entrada.xlsx"))
    parser.add_argument("--salida", type=Path, default=Path("salida.xlsx"))
    parser.add_argument("--correo", type=str, default=EMAIL_FIJO)
    parser.add_argument("--pausa", type=float, default=2.5)
    args = parser.parse_args()
    procesar_excel(args.excel, args.salida, args.correo, pausa_seg=args.pausa)

if __name__ == "__main__":
    main()
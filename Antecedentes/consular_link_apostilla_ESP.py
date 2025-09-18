import argparse
import random
import re
import time
from pathlib import Path
from typing import Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import sync_playwright, TimeoutError

# Configuración
BASE_URL = "https://tramites.cancilleria.gov.co/apostillalegalizacion/consulta/tramite.aspx"
EMAIL_FIJO = "apostillamen@gmail.com"
MSJ_NO_ENCONTRADO = "No se han encontrado registros"
MSJ_POSIBLE_ANTIBOT = "POSIBLE ANTIBOT/CAPTCHA"
MAX_REINTENTOS_CAPTCHA = 3


def limpiar_codigo(valor) -> Optional[str]:
    """Extrae solo dígitos y valida que inicie con '52'."""
    if pd.isna(valor) or not isinstance(valor, (str, int, float)):
        return None
    s = re.sub(r"\D", "", str(valor))
    if not s:
        return None
    if not s.startswith("52"):
        return None
    return s


def consultar_codigo(page, codigo: str, correo: str) -> Tuple[str, str]:
    """Consulta un código en el portal de la Cancillería con reintentos limitados."""
    print(f"  🔎 Buscando código {codigo} ...")

    sel_codigo = "#contenido_tbNumeroSolicitud"
    sel_correo = "#contenido_ucCorreoElectronico_tbCorreoElectronico"
    sel_btn = "#contenido_btnBuscar"
    sel_link_exito = "#contenido_ucInfor_lblMensajes2 a"
    sel_no_encontrado = "text=No se han encontrado registros"
    sel_error_captcha = "#contenido_validadorCaptcha"

    start_time = time.time()

    try:
        page.goto(BASE_URL, timeout=60_000)
        page.wait_for_selector(sel_codigo, timeout=30_000)
        page.fill(sel_codigo, codigo)
        page.wait_for_selector(sel_correo, timeout=30_000)
        page.fill(sel_correo, correo.strip())
    except Exception as e:
        return "ERROR", f"Error cargando página: {e}"

    for attempt in range(MAX_REINTENTOS_CAPTCHA):
        # Control de tiempo por fila (máx. 60s)
        if time.time() - start_time > 60:
            print("  ⏱️ Tiempo excedido en esta fila (60s).")
            return "TIMEOUT", "Tiempo máximo excedido"

        page.click(sel_btn)

        # 1. Intentar éxito
        try:
            link_handle = page.locator(sel_link_exito).first
            href = link_handle.get_attribute("href", timeout=5_000)
            if href:
                print(f"  ✅ Encontrado link: {href}")
                return "OK", href
        except TimeoutError:
            pass

        # 2. No encontrado
        if page.locator(sel_no_encontrado).count() > 0:
            print("  ❌ No encontrado")
            return "NO", MSJ_NO_ENCONTRADO

        # 3. CAPTCHA detectado
        if page.locator(sel_error_captcha).is_visible():
            print(f"  🤖 CAPTCHA detectado. Intento {attempt + 1}/{MAX_REINTENTOS_CAPTCHA}.")
            time.sleep(random.uniform(1, 2))
            continue

        time.sleep(2)

    print("  ⚠️ Posible antibot o sin respuesta tras varios intentos.")
    return "ANTIBOT", MSJ_POSIBLE_ANTIBOT


def procesar_excel(path_excel: Path, correo: str = EMAIL_FIJO, pausa_seg: float = 2.5):
    """Procesa el Excel de entrada y lo actualiza con los resultados."""
    wb = load_workbook(path_excel)
    ws = wb.active

    # Buscar índices de columnas
    headers = {cell.value.strip().upper(): cell.column_letter for cell in ws[1]}
    col_codigo = headers.get("CODIGO")
    col_link = headers.get("LINK")
    col_obs = headers.get("OBSERVACIONES")

    if not col_codigo or not col_link or not col_obs:
        raise ValueError("El Excel no contiene las columnas esperadas: CODIGO, LINK, OBSERVACIONES")

    rojo = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=100)
        page = browser.new_page()

        for row in range(2, ws.max_row + 1):
            codigo_raw = ws[f"{col_codigo}{row}"].value
            link_cell = ws[f"{col_link}{row}"]
            obs_cell = ws[f"{col_obs}{row}"]

            if link_cell.hyperlink:  # Ya tiene link
                print(f"\nFila {row}: {codigo_raw} (ya tiene link, se omite)")
                continue

            print(f"\nFila {row}: {codigo_raw}")
            codigo = limpiar_codigo(codigo_raw)
            if not codigo:
                obs_cell.value = "Código inválido o no inicia con 52"
                obs_cell.fill = rojo
                continue

            try:
                estado, valor = consultar_codigo(page, codigo, correo)
            except Exception as e:
                estado, valor = "ERROR", str(e)

            if estado == "OK":
                link_cell.value = "Link de apostilla"
                link_cell.hyperlink = valor
                link_cell.font = Font(color="0000FF", underline="single")
                obs_cell.value = "OK"
            elif estado == "NO":
                obs_cell.value = MSJ_NO_ENCONTRADO
                obs_cell.fill = rojo
            elif estado == "ANTIBOT":
                obs_cell.value = MSJ_POSIBLE_ANTIBOT
                obs_cell.fill = amarillo
            else:
                obs_cell.value = valor

            time.sleep(pausa_seg)

        browser.close()

    wb.save(path_excel)
    print(f"\n✅ Archivo actualizado: {path_excel.resolve()}")


def main():
    parser = argparse.ArgumentParser(description="Consulta de códigos de apostilla")
    parser.add_argument("--excel", type=Path, default=Path("entrada.xlsx"))
    parser.add_argument("--correo", type=str, default=EMAIL_FIJO)
    parser.add_argument("--pausa", type=float, default=2.5)
    args = parser.parse_args()

    procesar_excel(args.excel, args.correo, pausa_seg=args.pausa)


if __name__ == "__main__":
    main()
